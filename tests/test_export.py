"""export.py testleri - tmp_path'e gercek Excel dosyasi yazip openpyxl ile okuyarak dogrular."""

from __future__ import annotations

from openpyxl import load_workbook

from piyasa_komutani.data import PortfolioRow
from piyasa_komutani.export import HEADERS, ReportRow, write_report
from piyasa_komutani.scoring import ScoreResult


def _portfolio_row(symbol: str = "THYAO.IS") -> PortfolioRow:
    return PortfolioRow(symbol, "stock", 100.0, 285.50, "TRY")


def test_write_report_writes_headers_and_full_row(tmp_path) -> None:
    path = tmp_path / "sonuclar.xlsx"
    score = ScoreResult(score=2, recommendation="GUCLU AL", reasons=("RSI dusuk.", "MACD pozitif."))
    rows = [ReportRow(_portfolio_row(), 330.0, score)]

    write_report(rows, path)

    workbook = load_workbook(path)
    sheet = workbook.active
    header_values = [cell.value for cell in sheet[1]]
    data_values = [cell.value for cell in sheet[2]]

    assert header_values == HEADERS
    assert data_values == [
        "THYAO.IS",
        "stock",
        100.0,
        285.50,
        "TRY",
        330.0,
        2,
        "GUCLU AL",
        "RSI dusuk.; MACD pozitif.",
    ]


def test_write_report_writes_placeholder_for_missing_data(tmp_path) -> None:
    path = tmp_path / "sonuclar.xlsx"
    rows = [ReportRow(_portfolio_row("BILINMEYEN"), None, None)]

    write_report(rows, path)

    workbook = load_workbook(path)
    sheet = workbook.active
    data_values = [cell.value for cell in sheet[2]]

    # openpyxl bos string hucreleri geri okurken None olarak dondurur.
    assert data_values == ["BILINMEYEN", "stock", 100.0, 285.50, "TRY", "Veri yok", None, "VERI YOK", None]


def test_write_report_writes_reason_for_unavailable_score(tmp_path) -> None:
    path = tmp_path / "sonuclar.xlsx"
    score = ScoreResult(None, None, (), "Yetersiz gecmis veri: EMA26 icin en az 26 gun gerekli, mevcut 5 gun.")
    rows = [ReportRow(_portfolio_row(), 330.0, score)]

    write_report(rows, path)

    workbook = load_workbook(path)
    sheet = workbook.active
    data_values = [cell.value for cell in sheet[2]]

    assert data_values == [
        "THYAO.IS",
        "stock",
        100.0,
        285.50,
        "TRY",
        330.0,  # close fiyati hala gosteriliyor, sadece skor yok
        None,
        "Yetersiz gecmis veri: EMA26 icin en az 26 gun gerekli, mevcut 5 gun.",
        None,
    ]


def test_write_report_creates_missing_parent_directory(tmp_path) -> None:
    path = tmp_path / "output" / "nested" / "sonuclar.xlsx"
    rows = [ReportRow(_portfolio_row(), 330.0, ScoreResult(0, "NOTR", ()))]

    write_report(rows, path)

    assert path.exists()


def test_write_report_header_row_is_bold(tmp_path) -> None:
    path = tmp_path / "sonuclar.xlsx"
    write_report([], path)

    workbook = load_workbook(path)
    sheet = workbook.active

    assert all(cell.font.bold for cell in sheet[1])


def test_write_report_multiple_rows(tmp_path) -> None:
    path = tmp_path / "sonuclar.xlsx"
    rows = [
        ReportRow(_portfolio_row("AAA"), 10.0, ScoreResult(1, "AL", ("neden",))),
        ReportRow(_portfolio_row("BBB"), None, None),
        ReportRow(_portfolio_row("CCC"), 20.0, ScoreResult(-1, "SAT", ("neden",))),
    ]

    write_report(rows, path)

    workbook = load_workbook(path)
    sheet = workbook.active

    assert sheet.max_row == 4  # baslik + 3 satir
    assert [sheet.cell(row=r, column=1).value for r in range(2, 5)] == ["AAA", "BBB", "CCC"]
